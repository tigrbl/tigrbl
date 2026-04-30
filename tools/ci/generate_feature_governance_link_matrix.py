from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any

from common import repo_root


ROOT = repo_root()
REGISTRY_PATH = ROOT / ".ssot" / "registry.json"
DEFAULT_OUTPUT = ROOT / ".ssot" / "reports" / "feature-governance-link-matrix.md"
DEFAULT_CSV_OUTPUT = ROOT / ".ssot" / "reports" / "feature-governance-link-matrix.csv"
DEFAULT_XLSX_OUTPUT = ROOT / ".ssot" / "reports" / "feature-governance-link-matrix.xlsx"

HIGHLIGHTED_ADR_IDS = [
    "adr:1003",
    "adr:1008",
    "adr:1016",
    "adr:1024",
    "adr:1052",
    "adr:1088",
    "adr:1100",
    "adr:1120",
    "adr:1121",
]

HIGHLIGHTED_SPEC_IDS = [
    "spc:1000",
    "spc:2072",
    "spc:2087",
    "spc:2089",
    "spc:2090",
    "spc:2092",
    "spc:2105",
    "spc:2106",
    "spc:2107",
    "spc:2108",
    "spc:2110",
    "spc:2113",
    "spc:2138",
    "spc:2140",
]

SELECTED_FEATURE_IDS = [
    "feat:anon-access-rest-status-parity-001",
    "feat:app-framed-message-codec-runtime-001",
    "feat:atom-parity-dep-security-001",
    "feat:atom-parity-egress-asgi-send-001",
    "feat:atom-parity-egress-to-transport-response-001",
    "feat:atom-parity-ingress-transport-extract-001",
    "feat:atom-parity-resolve-assemble-001",
    "feat:atom-parity-response-error-to-transport-001",
    "feat:atom-parity-sys-handler-aggregate-001",
    "feat:binding-driven-ingress-001",
    "feat:bindingspec-event-subevent-schema-001",
    "feat:bindingspec-kernelplan-protocol-compilation-001",
    "feat:binding-subevent-phase-atom-legality-matrix-001",
    "feat:canonical-ingress-route-phase-cleanup-001",
    "feat:derived-runtime-subevent-taxonomy-001",
    "feat:dispatch-exchange-family-subevent-atoms-001",
    "feat:eventful-channel-state-metadata-001",
    "feat:eventkey-hook-bucket-compilation-001",
    "feat:eventstreamresponse-concrete-sse-class-001",
    "feat:executor-dispatch-removal-001",
    "feat:framing-decode-encode-atoms-001",
    "feat:http-jsonrpc-bindingspec-contract-001",
    "feat:http-rest-bindingspec-contract-001",
    "feat:http-rest-jsonrpc-atom-chains-001",
    "feat:http-stream-atom-chains-001",
    "feat:jsonrpc-20-runtime-surface-001",
    "feat:jsonrpc-batch-framing-001",
    "feat:jsonrpc-endpoint-key-001",
    "feat:jsonrpc-input-validation-before-persistence-001",
    "feat:jsonrpc-notification-204-projection-001",
    "feat:jsonrpc-persistence-error-sanitization-001",
    "feat:kernel-bootstrap-plan-parity-001",
    "feat:kernel-cache-invalidation-contract-001",
    "feat:kernelplan-dispatch-ownership-001",
    "feat:kernelz-mount-surface-001",
    "feat:kernelz-uix-surface-001",
    "feat:lifespan-runtime-chain-001",
    "feat:message-datagram-runtime-families-001",
    "feat:mutualtls-security-docs-runtime-alignment",
    "feat:operator-sse-event-stream-surface-001",
    "feat:operator-streaming-response-surface-001",
    "feat:operator-websocket-route-surface-001",
    "feat:phase-tree-error-branches-001",
    "feat:post-emit-completion-fence-compilation-001",
    "feat:protocol-anchor-ordering-parity-001",
    "feat:protocol-fused-segments-001",
    "feat:protocol-phase-tree-plans-001",
    "feat:protocol-runtime-boundary-certification-001",
    "feat:protocol-runtime-profile-pack-001",
    "feat:protocol-runtime-ssot-feature-granularity-001",
    "feat:protocol-runtime-test-evidence-suite-001",
    "feat:python-asgi-boundary-evidence-001",
    "feat:python-direct-runtime-microbench-001",
    "feat:python-runtime-ddl-initialization-boundary-001",
    "feat:python-runtime-performance-baseline-001",
    "feat:rest-create-success-status-001",
    "feat:rust-asgi-boundary-evidence-001",
    "feat:rust-direct-runtime-microbench-001",
    "feat:rust-protocol-plan-parity-001",
    "feat:rust-runtime-ddl-initialization-boundary-001",
    "feat:rust-runtime-performance-baseline-001",
    "feat:sse-event-framing-001",
    "feat:sse-lazy-iterator-runtime-001",
    "feat:sse-session-message-stream-chains-001",
    "feat:static-file-runtime-chain-001",
    "feat:subevent-transaction-units-001",
    "feat:transport-accept-emit-close-atoms-001",
    "feat:transport-bypass-removal-001",
    "feat:transport-dispatch-governance-001",
    "feat:transport-event-registry-001",
    "feat:uvicorn-protocol-mode-runtime-parity-001",
    "feat:uvicorn-rest-rpc-semantic-parity-001",
    "feat:websocket-concrete-runtime-class-001",
    "feat:websocket-wss-atom-chains-001",
    "feat:webtransport-bindingspec-contract-001",
    "feat:webtransport-transport-events-001",
    "feat:yield-iterator-producer-001",
]


def _load_registry(path: Path = REGISTRY_PATH) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _by_id(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {row["id"]: row for row in rows}


def _sorted_unique(values: list[str]) -> list[str]:
    return sorted(dict.fromkeys(values))


def _normalize_ids(ids: list[str], prefix: str) -> list[str]:
    return sorted(
        {
            value
            for value in ids
            if isinstance(value, str) and value.startswith(prefix)
        }
    )


def _normalized_path(path: str) -> str:
    return path.replace("\\", "/")


def _doc_number(doc_id: str) -> int:
    _, _, tail = doc_id.partition(":")
    return int(tail)


def _doc_filename(doc_id: str) -> str:
    prefix = "ADR" if doc_id.startswith("adr:") else "SPEC"
    return f"{prefix}-{_doc_number(doc_id):04d}"


def _validate_doc_entry(
    entry: dict[str, Any] | None,
    doc_id: str,
    *,
    root: Path,
    expected_prefix: str,
) -> list[str]:
    errors: list[str] = []
    if entry is None:
        return [f"Missing registry entry for {doc_id}."]

    path = entry.get("path")
    if not isinstance(path, str):
        return [f"Registry entry {doc_id} is missing a string path."]

    normalized = _normalized_path(path)
    expected_dir = f".ssot/{expected_prefix}/"
    if not normalized.startswith(expected_dir):
        errors.append(
            f"Registry entry {doc_id} points to {normalized}, expected it under {expected_dir}."
        )

    doc_path = root / normalized
    if not doc_path.exists():
        errors.append(f"Document file for {doc_id} does not exist: {normalized}.")

    expected_name = _doc_filename(doc_id)
    if expected_name not in Path(normalized).name:
        errors.append(
            f"Registry entry {doc_id} points to {normalized}, expected filename containing {expected_name}."
        )

    return errors


def build_report(
    registry: dict[str, Any],
    *,
    selected_feature_ids: list[str] | None = None,
    highlighted_adr_ids: list[str] | None = None,
    highlighted_spec_ids: list[str] | None = None,
    root: Path = ROOT,
) -> dict[str, Any]:
    selected_feature_ids = selected_feature_ids or SELECTED_FEATURE_IDS
    highlighted_adr_ids = highlighted_adr_ids or HIGHLIGHTED_ADR_IDS
    highlighted_spec_ids = highlighted_spec_ids or HIGHLIGHTED_SPEC_IDS

    features_by_id = _by_id(registry.get("features", []))
    adrs_by_id = _by_id(registry.get("adrs", []))
    specs_by_id = _by_id(registry.get("specs", []))

    errors: list[str] = []
    selected_features: list[dict[str, Any]] = []
    for feature_id in selected_feature_ids:
        feature = features_by_id.get(feature_id)
        if feature is None:
            errors.append(f"Missing selected feature {feature_id}.")
            continue
        if not feature_id.startswith("feat:"):
            errors.append(f"Selected row {feature_id} is not a feature id.")
            continue
        selected_features.append(feature)

    selected_features.sort(key=lambda item: item["id"])

    declared_adr_ids = _sorted_unique(
        [
            adr_id
            for feature in selected_features
            for adr_id in _normalize_ids(feature.get("adr_ids", []), "adr:")
        ]
    )
    declared_spec_ids = _sorted_unique(
        [
            spec_id
            for feature in selected_features
            for spec_id in _normalize_ids(feature.get("spec_ids", []), "spc:")
        ]
    )

    highlighted_adr_ids = _sorted_unique(list(highlighted_adr_ids))
    highlighted_spec_ids = _sorted_unique(list(highlighted_spec_ids))
    additional_adr_ids = [doc_id for doc_id in declared_adr_ids if doc_id not in highlighted_adr_ids]
    additional_spec_ids = [doc_id for doc_id in declared_spec_ids if doc_id not in highlighted_spec_ids]

    column_adr_ids = highlighted_adr_ids + additional_adr_ids
    column_spec_ids = highlighted_spec_ids + additional_spec_ids

    for doc_id in column_adr_ids:
        errors.extend(
            _validate_doc_entry(adrs_by_id.get(doc_id), doc_id, root=root, expected_prefix="adr")
        )
    for doc_id in column_spec_ids:
        errors.extend(
            _validate_doc_entry(specs_by_id.get(doc_id), doc_id, root=root, expected_prefix="specs")
        )

    if errors:
        raise ValueError("\n".join(errors))

    rows = [
        {
            "feature_id": feature["id"],
            "title": str(feature.get("title", "")),
            "implementation_status": str(feature.get("implementation_status", "")),
            "adr_ids": _normalize_ids(feature.get("adr_ids", []), "adr:"),
            "spec_ids": _normalize_ids(feature.get("spec_ids", []), "spc:"),
        }
        for feature in selected_features
    ]

    return {
        "source_registry": ".ssot/registry.json",
        "row_selection_rule": (
            "Explicit runtime/protocol feature seed set from the prior inventory, filtered to real feat:* rows, "
            "with stable row ordering by feature_id."
        ),
        "column_selection_rule": (
            "Highlighted governance IDs first, then any additional ADR/SPEC IDs directly declared by the selected "
            "features through adr_ids or spec_ids."
        ),
        "cell_semantics": "X means the feature directly declares that ADR or SPEC in its own adr_ids or spec_ids.",
        "rows": rows,
        "highlighted_adr_ids": highlighted_adr_ids,
        "highlighted_spec_ids": highlighted_spec_ids,
        "additional_adr_ids": additional_adr_ids,
        "additional_spec_ids": additional_spec_ids,
        "column_adr_ids": column_adr_ids,
        "column_spec_ids": column_spec_ids,
        "legend": {
            "adr": {
                doc_id: str(adrs_by_id[doc_id].get("title", ""))
                for doc_id in column_adr_ids
            },
            "spec": {
                doc_id: str(specs_by_id[doc_id].get("title", ""))
                for doc_id in column_spec_ids
            },
        },
    }


def _escape_cell(value: str) -> str:
    return value.replace("|", "\\|").replace("\n", " ").strip()


def _render_table(
    rows: list[dict[str, Any]],
    doc_ids: list[str],
    *,
    doc_key: str,
) -> list[str]:
    headers = ["feature_id", "title", "implementation_status", *doc_ids]
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        linked = set(row[doc_key])
        cells = [
            f"`{row['feature_id']}`",
            _escape_cell(row["title"]),
            f"`{row['implementation_status']}`",
            *["X" if doc_id in linked else "" for doc_id in doc_ids],
        ]
        lines.append("| " + " | ".join(cells) + " |")
    return lines


def render_markdown(report: dict[str, Any]) -> str:
    lines = [
        "# Feature-to-Governance Link Matrix",
        "",
        "This report is generated from `.ssot/registry.json`.",
        "",
        "- `source file`: `.ssot/registry.json`",
        f"- `row-selection rule`: {report['row_selection_rule']}",
        f"- `column-selection rule`: {report['column_selection_rule']}",
        f"- `cell semantics`: {report['cell_semantics']}",
        "",
        "The matrix is one logical report rendered as separate ADR and SPEC tables so the Markdown remains readable.",
        "",
        "## ADR Matrix",
        "",
        "Highlighted ADRs appear first, followed by any additional ADRs directly declared by the selected features.",
        "",
    ]
    lines.extend(_render_table(report["rows"], report["column_adr_ids"], doc_key="adr_ids"))
    lines.extend(
        [
            "",
            "## SPEC Matrix",
            "",
            "Highlighted SPECs appear first, followed by any additional SPECs directly declared by the selected features.",
            "",
        ]
    )
    lines.extend(_render_table(report["rows"], report["column_spec_ids"], doc_key="spec_ids"))
    lines.extend(
        [
            "",
            "## Legend",
            "",
            "### Highlighted ADRs/SPECs",
            "",
            "| ID | Title |",
            "|---|---|",
        ]
    )
    for doc_id in report["highlighted_adr_ids"]:
        lines.append(f"| `{doc_id}` | {_escape_cell(report['legend']['adr'][doc_id])} |")
    for doc_id in report["highlighted_spec_ids"]:
        lines.append(f"| `{doc_id}` | {_escape_cell(report['legend']['spec'][doc_id])} |")
    lines.extend(
        [
            "",
            "### Additional Declared ADRs/SPECs",
            "",
            "| ID | Title |",
            "|---|---|",
        ]
    )
    additional_ids = report["additional_adr_ids"] + report["additional_spec_ids"]
    if additional_ids:
        for doc_id in report["additional_adr_ids"]:
            lines.append(f"| `{doc_id}` | {_escape_cell(report['legend']['adr'][doc_id])} |")
        for doc_id in report["additional_spec_ids"]:
            lines.append(f"| `{doc_id}` | {_escape_cell(report['legend']['spec'][doc_id])} |")
    else:
        lines.append("| none | No additional ADR or SPEC IDs were directly declared by the selected features beyond the highlighted set. |")
    return "\n".join(lines) + "\n"


def render_csv(report: dict[str, Any]) -> str:
    headers = [
        "feature_id",
        "title",
        "implementation_status",
        *report["column_adr_ids"],
        *report["column_spec_ids"],
    ]
    rows: list[list[str]] = [headers]
    for row in report["rows"]:
        linked_adrs = set(row["adr_ids"])
        linked_specs = set(row["spec_ids"])
        rows.append(
            [
                row["feature_id"],
                row["title"],
                row["implementation_status"],
                *["X" if doc_id in linked_adrs else "" for doc_id in report["column_adr_ids"]],
                *["X" if doc_id in linked_specs else "" for doc_id in report["column_spec_ids"]],
            ]
        )

    from io import StringIO

    buffer = StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerows(rows)
    return buffer.getvalue()


def write_xlsx(report: dict[str, Any], output_path: Path) -> None:
    try:
        from openpyxl import Workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "XLSX output requires openpyxl. Run this generator with `uv run python` or install openpyxl."
        ) from exc

    workbook = Workbook()
    adr_sheet = workbook.active
    adr_sheet.title = "ADR Matrix"
    spec_sheet = workbook.create_sheet("SPEC Matrix")
    legend_sheet = workbook.create_sheet("Legend")

    adr_headers = ["feature_id", "title", "implementation_status", *report["column_adr_ids"]]
    spec_headers = ["feature_id", "title", "implementation_status", *report["column_spec_ids"]]

    adr_sheet.append(adr_headers)
    spec_sheet.append(spec_headers)

    for row in report["rows"]:
        linked_adrs = set(row["adr_ids"])
        linked_specs = set(row["spec_ids"])
        adr_sheet.append(
            [
                row["feature_id"],
                row["title"],
                row["implementation_status"],
                *["X" if doc_id in linked_adrs else "" for doc_id in report["column_adr_ids"]],
            ]
        )
        spec_sheet.append(
            [
                row["feature_id"],
                row["title"],
                row["implementation_status"],
                *["X" if doc_id in linked_specs else "" for doc_id in report["column_spec_ids"]],
            ]
        )

    legend_sheet.append(["group", "id", "title"])
    for doc_id in report["highlighted_adr_ids"]:
        legend_sheet.append(["highlighted_adr", doc_id, report["legend"]["adr"][doc_id]])
    for doc_id in report["highlighted_spec_ids"]:
        legend_sheet.append(["highlighted_spec", doc_id, report["legend"]["spec"][doc_id]])
    for doc_id in report["additional_adr_ids"]:
        legend_sheet.append(["additional_adr", doc_id, report["legend"]["adr"][doc_id]])
    for doc_id in report["additional_spec_ids"]:
        legend_sheet.append(["additional_spec", doc_id, report["legend"]["spec"][doc_id]])

    for sheet in (adr_sheet, spec_sheet, legend_sheet):
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 48)
        sheet.freeze_panes = "D2" if sheet.title != "Legend" else "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_report(
    report: dict[str, Any],
    markdown: str,
    csv_text: str,
    markdown_path: Path,
    csv_path: Path,
    xlsx_path: Path | None = None,
) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    markdown_path.parent.mkdir(parents=True, exist_ok=True)
    markdown_path.write_text(markdown, encoding="utf-8")
    csv_path.write_text(csv_text, encoding="utf-8")
    if xlsx_path is not None:
        write_xlsx(report, xlsx_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--csv-output", type=Path, default=DEFAULT_CSV_OUTPUT)
    parser.add_argument("--xlsx-output", type=Path, default=DEFAULT_XLSX_OUTPUT)
    args = parser.parse_args()

    report = build_report(_load_registry())
    write_report(
        report,
        render_markdown(report),
        render_csv(report),
        args.output,
        args.csv_output,
        args.xlsx_output,
    )


if __name__ == "__main__":
    main()
